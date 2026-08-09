#!/usr/bin/env python3
"""指定アドレスからの当日分メールを抽出し、要約用JSONを出力する読み取り専用スクリプト。

対象アドレスは mails.xlsx の「欲しい処理」列を参照し、
【就活】または【企業】で始まる行のメールアドレスのみを対象とする。
削除・既読化・変更は一切行わない(gmail.readonlyスコープのみ使用)。
"""

import argparse
import json
import sys
from datetime import date, datetime, timedelta, timezone
from email.utils import parseaddr, parsedate_to_datetime
from pathlib import Path

import openpyxl
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from gmail_utils import extract_body, get_credentials, get_header

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / "mails.xlsx"
MAIL_DATA_DIR = BASE_DIR / "mail_data"

JST = timezone(timedelta(hours=9))
RULE_PREFIXES = {"【就活】": "job", "【企業】": "company"}


def load_target_addresses() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]

    targets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        comment = (d.get("欲しい処理") or "").strip()
        addr = (d.get("メールアドレス") or "").strip().lower()
        if not addr:
            continue
        for prefix, rule_type in RULE_PREFIXES.items():
            if comment.startswith(prefix):
                targets[addr] = {
                    "type": rule_type,
                    "comment": comment,
                    "name": d.get("表示名") or "",
                    "base_tag": (d.get("Base Tag") or "").strip(),
                }
                break
    return targets


def fetch_candidates(service, target_date: date) -> list[dict]:
    after_date = (target_date - timedelta(days=1)).strftime("%Y/%m/%d")
    resp = (
        service.users()
        .messages()
        .list(userId="me", q=f"after:{after_date}", maxResults=200)
        .execute()
    )
    messages = resp.get("messages", [])

    results = []
    for m in messages:
        msg = service.users().messages().get(userId="me", id=m["id"], format="full").execute()
        headers = msg["payload"].get("headers", [])
        results.append(
            {
                "id": msg["id"],
                "from": get_header(headers, "From"),
                "subject": get_header(headers, "Subject"),
                "date": get_header(headers, "Date"),
                "body": extract_body(msg["payload"]),
            }
        )
    return results


def main():
    parser = argparse.ArgumentParser(description="対象アドレスからの当日分メールを抽出しJSON出力する")
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="対象日(YYYY-MM-DD、JST)。省略時は実行時点のJSTでの今日。",
    )
    args = parser.parse_args()

    target_date = (
        datetime.strptime(args.date, "%Y-%m-%d").date()
        if args.date
        else datetime.now(JST).date()
    )

    targets = load_target_addresses()
    if not targets:
        sys.exit("mails.xlsx に【就活】/【企業】対象アドレスが見つかりません。")

    creds = get_credentials()
    service = build("gmail", "v1", credentials=creds)

    try:
        candidates = fetch_candidates(service, target_date)
    except HttpError as error:
        sys.exit(f"Gmail API呼び出し中にエラーが発生しました: {error}")

    matched = []
    for c in candidates:
        _, addr = parseaddr(c["from"])
        addr = addr.lower()
        if addr not in targets:
            continue

        try:
            dt = parsedate_to_datetime(c["date"])
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            continue
        if dt.astimezone(JST).date() != target_date:
            continue

        rule = targets[addr]
        matched.append(
            {
                "id": c["id"],
                "from_address": addr,
                "from_name": rule["name"] or c["from"],
                "subject": c["subject"],
                "date_jst": dt.astimezone(JST).isoformat(),
                "body": c["body"],
                "rule_type": rule["type"],
                "rule_comment": rule["comment"],
                "base_tag": rule["base_tag"],
            }
        )

    output = {
        "target_date": target_date.isoformat(),
        "generated_at": datetime.now(JST).isoformat(),
        "count": len(matched),
        "emails": matched,
    }
    MAIL_DATA_DIR.mkdir(exist_ok=True)
    output_path = MAIL_DATA_DIR / f"{target_date.isoformat()}.json"
    output_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"対象メール {len(matched)} 件を {output_path} に保存しました。(対象日: {target_date})")


if __name__ == "__main__":
    main()
